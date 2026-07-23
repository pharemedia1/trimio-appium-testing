#!/usr/bin/env python3
"""Build docs/Trimio-Test-Cases.xlsx from the case tables in scripts/testcases/.

    python3 scripts/generate_test_cases.py [output.xlsx]

The workbook is a build output: edit the Python case tables, never the .xlsx, so the
suite stays reviewable in diffs. Requires openpyxl (`pip3 install --user openpyxl`).
"""

import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from testcases import COLUMNS
from testcases import admin, auth, client, crosscutting, professional, store, web

MODULES = [auth, client, professional, admin, store, web, crosscutting]

# ---- styling ---------------------------------------------------------------
BRAND = "0066B2"          # Trimio blue (matches AdminUI.brand)
BRAND_DARK = "0C1B2E"     # portal sidebar navy
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
TITLE_FILL = PatternFill("solid", fgColor=BRAND_DARK)
ZEBRA_FILL = PatternFill("solid", fgColor="F4F7FB")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
BODY_FONT = Font(size=10)
THIN = Side(style="thin", color="D6DEE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PRIORITY_FILL = {
    "P1": PatternFill("solid", fgColor="FFD9D9"),
    "P2": PatternFill("solid", fgColor="FFF2CC"),
    "P3": PatternFill("solid", fgColor="E2EFDA"),
}
AUTO_FILL = {
    "Automated": PatternFill("solid", fgColor="D9EAD3"),
    "Manual": PatternFill("solid", fgColor="EDEDED"),
}


def _style_header(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def write_case_sheet(wb, module):
    ws = wb.create_sheet(module.SHEET)
    ws.append([name for name, _ in COLUMNS])
    _style_header(ws, len(COLUMNS))

    for i, case in enumerate(module.CASES, 1):
        tc_id = f"{module.PREFIX}-{i:03d}"
        ws.append([
            tc_id, case.module, case.sub, case.role, case.platform, case.title,
            case.pre, case.numbered_steps(), case.data, case.exp,
            case.pri, case.typ, case.auto, case.script,
        ])
        row = ws.max_row
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if i % 2 == 0:
                cell.fill = ZEBRA_FILL
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=11).fill = PRIORITY_FILL.get(case.pri, ZEBRA_FILL)
        ws.cell(row=row, column=11).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=13).fill = AUTO_FILL.get(case.auto, ZEBRA_FILL)
        ws.cell(row=row, column=13).alignment = Alignment(horizontal="center", vertical="top")

    for idx, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    return ws


def write_cover(wb, total, by_priority, by_platform, by_auto, by_type):
    ws = wb.create_sheet("00-Cover", 0)
    ws.merge_cells("A1:D1")
    ws["A1"] = "Trimio — Master Test Case Suite"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 40

    rows = [
        ("", ""),
        ("Scope", "Client · Professional · Admin · Vendor · Support — mobile app, web portal and Store"),
        ("Surfaces", "Flutter Android/iOS app; Flutter Web staff portal (admin + vendor); Node/Express API"),
        ("Automation", "Appium + TestNG (mobile) · Playwright Java + TestNG (web) — project TrimioAutomation"),
        ("Generated from", "scripts/testcases/*.py via scripts/generate_test_cases.py"),
        ("Analysis", "docs/Trimio-Test-Analysis.md"),
        ("", ""),
        ("Total test cases", total),
    ]
    for label, value in rows:
        ws.append([label, value])
        r = ws.max_row
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    def block(title, counter, order=None):
        ws.append([])
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12, color=BRAND)
        keys = order or sorted(counter, key=lambda k: -counter[k])
        for key in keys:
            if counter.get(key):
                ws.append([key, counter[key]])
                ws.cell(row=ws.max_row, column=1).font = BODY_FONT

    block("By priority", by_priority, order=["P1", "P2", "P3"])
    block("By platform", by_platform)
    block("By automation status", by_auto)
    block("By test type", by_type)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    return ws


def write_summary(wb, counts):
    ws = wb.create_sheet("08-Coverage-Summary")
    ws.append(["Worksheet", "Cases", "P1", "P2", "P3", "Automated", "Manual"])
    _style_header(ws, 7)
    for sheet, c in counts:
        ws.append([sheet, c["total"], c["P1"], c["P2"], c["P3"], c["Automated"], c["Manual"]])
        for col in range(1, 8):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = BODY_FONT
            cell.border = BORDER
    totals = ["TOTAL"] + [sum(c[k] for _, c in counts)
                          for k in ("total", "P1", "P2", "P3", "Automated", "Manual")]
    ws.append(totals)
    for col in range(1, 8):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True, size=10)
        cell.border = BORDER
    ws.column_dimensions["A"].width = 30
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 12
    try:
        ws.add_table(Table(displayName="Coverage", ref=f"A1:G{ws.max_row}",
                           tableStyleInfo=TableStyleInfo(name="TableStyleLight9", showRowStripes=True)))
    except ValueError:
        pass  # table styling is cosmetic
    return ws


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else \
        Path(__file__).resolve().parent.parent / "docs" / "Trimio-Test-Cases.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    by_priority, by_platform, by_auto, by_type = Counter(), Counter(), Counter(), Counter()
    counts = []

    for module in MODULES:
        write_case_sheet(wb, module)
        c = Counter()
        for case in module.CASES:
            c["total"] += 1
            c[case.pri] += 1
            c[case.auto] += 1
            by_priority[case.pri] += 1
            by_platform[case.platform] += 1
            by_auto[case.auto] += 1
            by_type[case.typ] += 1
        counts.append((module.SHEET, c))

    total = sum(c["total"] for _, c in counts)
    write_cover(wb, total, by_priority, by_platform, by_auto, by_type)
    write_summary(wb, counts)

    wb.save(out)
    print(f"Wrote {out} — {total} test cases across {len(MODULES)} module sheets")
    for sheet, c in counts:
        print(f"  {sheet:<26} {c['total']:>4} cases  "
              f"(P1={c['P1']:>3} P2={c['P2']:>3} P3={c['P3']:>3}, automated={c['Automated']:>3})")


if __name__ == "__main__":
    main()
